import csv
from io import BytesIO
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

# ----------------- CSV Generator -----------------
def generate_csv_report(headers, rows):
    """
    Generates a CSV report as bytes.
    """
    buffer = BytesIO()
    # Write UTF-8 BOM for Excel compatibility
    buffer.write(b'\xef\xbb\xbf')
    
    import io
    string_io = io.StringIO()
    csv_writer = csv.writer(string_io)
    csv_writer.writerow(headers)
    csv_writer.writerows(rows)
    
    buffer.write(string_io.getvalue().encode('utf-8'))
    buffer.seek(0)
    return buffer.getvalue()


# ----------------- Excel Generator -----------------
def generate_excel_report(sheet_name, headers, rows):
    """
    Generates a beautifully styled Excel spreadsheet using openpyxl.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:30] # Excel limit is 31 chars
    
    # Enable gridlines
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    title_font = Font(name='Arial', size=16, bold=True, color='1A365D')
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2B6CB0', end_color='2B6CB0', fill_type='solid')
    data_font = Font(name='Arial', size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E0'),
        right=Side(style='thin', color='CBD5E0'),
        top=Side(style='thin', color='CBD5E0'),
        bottom=Side(style='thin', color='CBD5E0')
    )

    # 1. Add Title Block
    ws.append([sheet_name.upper()])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1).font = title_font
    ws.row_dimensions[1].height = 30
    ws.append([]) # Empty spacer row
    
    # 2. Add Headers
    ws.append(headers)
    header_row_idx = 3
    ws.row_dimensions[header_row_idx].height = 24
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # 3. Add Rows
    for row_idx, row_data in enumerate(rows, start=4):
        ws.row_dimensions[row_idx].height = 20
        # Convert decimal fields to float or string, check cell formatting
        converted_row = []
        for val in row_data:
            if isinstance(val, Decimal):
                converted_row.append(float(val))
            else:
                converted_row.append(val)
                
        ws.append(converted_row)
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            
            # Formatting floats
            val = cell.value
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal='right', vertical='center')
                if isinstance(val, float):
                    cell.number_format = '$#,##0.00'
            else:
                cell.alignment = left_align

    # 4. Auto-fit Column Widths
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1: # Ignore title block in width calculation
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ----------------- PDF Generator -----------------
def generate_pdf_report(title, headers, rows):
    """
    Generates a clean, readable PDF report in landscape format using reportlab.
    """
    buffer = BytesIO()
    # Using landscape format for tabular reports to prevent clipping
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(letter),
        rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30
    )
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'PdfReportTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#1A365D'),
        alignment=1 # Center
    )
    
    header_style = ParagraphStyle(
        'PdfReportHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.white,
        alignment=1
    )
    
    cell_style = ParagraphStyle(
        'PdfReportCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        textColor=colors.HexColor('#2D3748')
    )

    elements = []
    
    # Title
    elements.append(Paragraph(title.upper(), title_style))
    elements.append(Spacer(1, 15))
    
    # Format Table Data
    table_data = []
    # Add Header Paragraphs
    table_data.append([Paragraph(h, header_style) for h in headers])
    
    # Add Row Paragraphs
    for r in rows:
        formatted_row = []
        for val in r:
            if isinstance(val, Decimal) or isinstance(val, float):
                txt = f"${val:,.2f}"
            elif val is None:
                txt = "N/A"
            else:
                txt = str(val)
            formatted_row.append(Paragraph(txt, cell_style))
        table_data.append(formatted_row)

    # Determine Column Widths (spread equally across page width ~ 730 points for letter landscape)
    num_cols = len(headers)
    col_width = 730 / num_cols
    
    report_table = Table(table_data, colWidths=[col_width] * num_cols)
    
    # Styling Table
    report_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2B6CB0')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E0')),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#1A365D')),
        # Alternate row backgrounds
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F7FAFC')]),
    ]))
    
    elements.append(report_table)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()
