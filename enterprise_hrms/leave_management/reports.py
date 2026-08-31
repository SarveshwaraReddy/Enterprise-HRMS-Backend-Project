import csv
import io
import datetime
from django.http import HttpResponse

# ReportLab imports for PDF generation
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# openpyxl import for Excel generation
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from enterprise_hrms.departments.models import Department
from .models import LeaveRequest, LeaveBalance, LeaveType


def generate_leave_history_pdf(queryset, title="Employee Leave History Report"):
    """
    Generates a PDF response containing leave history records.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#1E293B'),
        alignment=1, # Center
        spaceAfter=12
    )
    meta_style = ParagraphStyle(
        'MetaStyle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#64748B'),
        alignment=1,
        spaceAfter=18
    )

    story = [
        Paragraph(title, title_style),
        Paragraph(f"Generated on {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | HRMS Enterprise Portal", meta_style),
        Spacer(1, 10)
    ]

    table_data = [
        ["ID", "Employee", "Department", "Leave Type", "Start Date", "End Date", "Days", "Status", "Applied At"]
    ]

    cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=8, leading=10)

    for req in queryset:
        table_data.append([
            str(req.id),
            Paragraph(f"{req.employee.first_name} {req.employee.last_name}", cell_style),
            Paragraph(req.employee.department.name if req.employee.department else "-", cell_style),
            Paragraph(req.leave_type.name, cell_style),
            str(req.start_date),
            str(req.end_date),
            str(req.total_days),
            req.get_status_display(),
            req.applied_at.strftime('%Y-%m-%d')
        ])

    col_widths = [30, 100, 90, 80, 65, 65, 35, 90, 65]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))

    story.append(t)
    doc.build(story)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Leave_History_{datetime.date.today()}.pdf"'
    return response


def generate_annual_leave_register_excel(year=None):
    """
    Generates an Excel workbook (.xlsx) representing the Annual Leave Register.
    """
    if year is None:
        year = datetime.date.today().year

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Leave Register {year}"
    ws.views.sheetView[0].showGridLines = True

    # Title Banner
    ws.merge_cells("A1:H1")
    ws["A1"] = f"Enterprise HRMS - Annual Leave Register ({year})"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    headers = [
        "Employee ID", "Employee Name", "Department", "Leave Type",
        "Allocated Days", "Used Days", "Remaining Days", "Paid/Unpaid"
    ]
    ws.append([]) # Blank row
    ws.append(headers)

    header_row = 3
    ws.row_dimensions[header_row].height = 25
    header_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    balances = LeaveBalance.objects.filter(year=year).select_related('employee', 'employee__department', 'leave_type')

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    row_idx = header_row + 1
    for b in balances:
        ws.append([
            b.employee.employee_id,
            f"{b.employee.first_name} {b.employee.last_name}",
            b.employee.department.name if b.employee.department else "-",
            b.leave_type.name,
            b.allocated_days,
            b.used_days,
            b.remaining_days,
            "Paid" if b.leave_type.is_paid else "Unpaid"
        ])

        for col in range(1, 9):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row_idx += 1

    # Adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Annual_Leave_Register_{year}.xlsx"'
    return response


def generate_department_leave_summary_excel(year=None):
    """
    Generates an Excel workbook (.xlsx) summarizing leave utilization by department.
    """
    if year is None:
        year = datetime.date.today().year

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Dept Summary {year}"
    ws.views.sheetView[0].showGridLines = True

    ws.merge_cells("A1:E1")
    ws["A1"] = f"Department Leave Summary ({year})"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    headers = ["Department Code", "Department Name", "Total Leave Requests", "Approved Days Taken", "Pending Requests"]
    ws.append([])
    ws.append(headers)

    header_row = 3
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    depts = Department.objects.all()
    row_idx = header_row + 1
    for dept in depts:
        dept_requests = LeaveRequest.objects.filter(employee__department=dept, start_date__year=year)
        total_count = dept_requests.count()
        approved_days = sum(r.total_days for r in dept_requests.filter(status='approved'))
        pending_count = dept_requests.filter(status__startswith='pending').count()

        ws.append([dept.code, dept.name, total_count, approved_days, pending_count])
        row_idx += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Department_Leave_Summary_{year}.xlsx"'
    return response


def generate_leave_transactions_csv(queryset):
    """
    Generates a CSV HttpResponse of leave transactions.
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="Leave_Transactions_{datetime.date.today()}.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'Transaction ID', 'Employee ID', 'Employee Name', 'Department',
        'Leave Type', 'Start Date', 'End Date', 'Total Days', 'Status', 'Reason', 'Applied At'
    ])

    for req in queryset:
        writer.writerow([
            req.id,
            req.employee.employee_id,
            f"{req.employee.first_name} {req.employee.last_name}",
            req.employee.department.name if req.employee.department else '',
            req.leave_type.code,
            req.start_date,
            req.end_date,
            req.total_days,
            req.status,
            req.reason,
            req.applied_at.strftime('%Y-%m-%d %H:%M:%S')
        ])

    return response
