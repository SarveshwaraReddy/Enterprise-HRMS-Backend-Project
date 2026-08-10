import csv
import io
import datetime
from django.http import HttpResponse
from django.db.models import Avg, F, ExpressionWrapper, fields as db_fields

# ReportLab imports for PDF generation
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# openpyxl import for Excel generation
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import Asset, AssetAssignment, SupportTicket, SoftwareLicense


# ─────────────────────────────────────────────
# Shared helpers
# ─────────────────────────────────────────────

def _pdf_doc(buffer, title, subtitle=None):
    """Build a ReportLab SimpleDocTemplate and return (doc, story)."""
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=36, leftMargin=36,
        topMargin=36, bottomMargin=36,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title', parent=styles['Heading1'],
        fontSize=18, leading=22,
        textColor=colors.HexColor('#1E293B'),
        alignment=1, spaceAfter=6,
    )
    meta_style = ParagraphStyle(
        'Meta', parent=styles['Normal'],
        fontSize=9, textColor=colors.HexColor('#64748B'),
        alignment=1, spaceAfter=18,
    )
    story = [Paragraph(title, title_style)]
    if subtitle:
        story.append(Paragraph(subtitle, meta_style))
    story.append(
        Paragraph(
            f"Generated on {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Enterprise HRMS",
            meta_style,
        )
    )
    story.append(Spacer(1, 10))
    return doc, story


def _table_style():
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ])


def _excel_header(ws, title, headers):
    """Apply standard header styling to an Excel worksheet."""
    col_count = len(headers)
    ws.merge_cells(f"A1:{openpyxl.utils.get_column_letter(col_count)}1")
    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.append([])   # blank row
    ws.append(headers)

    hdr_row = 3
    for col in range(1, col_count + 1):
        cell = ws.cell(row=hdr_row, column=col)
        cell.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[hdr_row].height = 22
    return hdr_row + 1   # first data row index


def _thin_border():
    side = Side(style='thin', color='E2E8F0')
    return Border(left=side, right=side, top=side, bottom=side)


def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=8)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max(max_len + 4, 12)


# ─────────────────────────────────────────────
# Asset Report
# ─────────────────────────────────────────────

def generate_asset_report_pdf(queryset=None):
    """Generate a PDF asset inventory report."""
    if queryset is None:
        queryset = Asset.objects.select_related('category').all()

    buffer = io.BytesIO()
    doc, story = _pdf_doc(buffer, "Asset Inventory Report")

    cell_style = getSampleStyleSheet()['Normal']
    cell_style.fontSize = 7

    headers = ["Asset Code", "Name", "Category", "Serial No.", "Vendor",
               "Status", "Location", "Purchase Date", "Warranty Expiry"]
    data = [headers]
    for a in queryset:
        data.append([
            a.asset_code, a.name,
            a.category.name if a.category else '-',
            a.serial_number or '-',
            a.vendor or '-',
            a.get_status_display(),
            a.location or '-',
            str(a.purchase_date) if a.purchase_date else '-',
            str(a.warranty_expiry_date) if a.warranty_expiry_date else '-',
        ])

    t = Table(data, repeatRows=1)
    t.setStyle(_table_style())
    story.append(t)
    doc.build(story)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="Asset_Report_{datetime.date.today()}.pdf"'
    )
    return response


def generate_asset_report_excel(queryset=None):
    """Generate an Excel asset inventory report."""
    if queryset is None:
        queryset = Asset.objects.select_related('category').all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asset Inventory"
    headers = [
        "Asset Code", "Name", "Category", "Serial No.", "Vendor",
        "Status", "Location", "Purchase Date", "Warranty Expiry", "Purchase Cost",
    ]
    row_idx = _excel_header(ws, "Enterprise HRMS – Asset Inventory Report", headers)
    border = _thin_border()

    for a in queryset:
        row = [
            a.asset_code, a.name,
            a.category.name if a.category else '',
            a.serial_number or '',
            a.vendor or '',
            a.get_status_display(),
            a.location or '',
            str(a.purchase_date) if a.purchase_date else '',
            str(a.warranty_expiry_date) if a.warranty_expiry_date else '',
            float(a.purchase_cost) if a.purchase_cost else '',
        ]
        ws.append(row)
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = border
            ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center")
        row_idx += 1

    _autofit(ws)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="Asset_Report_{datetime.date.today()}.xlsx"'
    )
    return response


def generate_asset_report_csv(queryset=None):
    """Generate a CSV asset inventory report."""
    if queryset is None:
        queryset = Asset.objects.select_related('category').all()

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = (
        f'attachment; filename="Asset_Report_{datetime.date.today()}.csv"'
    )
    writer = csv.writer(response)
    writer.writerow([
        "Asset Code", "Name", "Category", "Serial No.", "Vendor",
        "Status", "Location", "Purchase Date", "Warranty Expiry", "Purchase Cost",
    ])
    for a in queryset:
        writer.writerow([
            a.asset_code, a.name,
            a.category.name if a.category else '',
            a.serial_number or '',
            a.vendor or '',
            a.get_status_display(),
            a.location or '',
            a.purchase_date or '',
            a.warranty_expiry_date or '',
            a.purchase_cost or '',
        ])
    return response


# ─────────────────────────────────────────────
# IT Support Report
# ─────────────────────────────────────────────

def generate_support_report_pdf(queryset=None):
    """Generate a PDF IT support ticket report."""
    if queryset is None:
        queryset = SupportTicket.objects.select_related('employee', 'assigned_engineer').all()

    buffer = io.BytesIO()
    doc, story = _pdf_doc(buffer, "IT Support Ticket Report")

    headers = ["Ticket No.", "Employee", "Priority", "Category", "Status",
               "Subject", "Assigned To", "Created At", "Closed At"]
    data = [headers]
    for t in queryset:
        data.append([
            t.ticket_number,
            f"{t.employee.first_name} {t.employee.last_name}",
            t.get_priority_display(),
            t.get_category_display(),
            t.get_status_display(),
            t.subject[:40],
            f"{t.assigned_engineer.first_name} {t.assigned_engineer.last_name}" if t.assigned_engineer else '-',
            t.created_at.strftime('%Y-%m-%d'),
            t.closed_at.strftime('%Y-%m-%d') if t.closed_at else '-',
        ])

    tbl = Table(data, repeatRows=1)
    tbl.setStyle(_table_style())
    story.append(tbl)
    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="Support_Report_{datetime.date.today()}.pdf"'
    )
    return response


def generate_support_report_excel(queryset=None):
    """Generate an Excel IT support ticket report."""
    if queryset is None:
        queryset = SupportTicket.objects.select_related('employee', 'assigned_engineer').all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Support Tickets"
    headers = [
        "Ticket No.", "Employee", "Priority", "Category",
        "Status", "Subject", "Assigned Engineer", "Created At", "Closed At",
    ]
    row_idx = _excel_header(ws, "Enterprise HRMS – IT Support Report", headers)
    border = _thin_border()

    for t in queryset:
        row = [
            t.ticket_number,
            f"{t.employee.first_name} {t.employee.last_name}",
            t.get_priority_display(),
            t.get_category_display(),
            t.get_status_display(),
            t.subject,
            f"{t.assigned_engineer.first_name} {t.assigned_engineer.last_name}" if t.assigned_engineer else '',
            t.created_at.strftime('%Y-%m-%d %H:%M'),
            t.closed_at.strftime('%Y-%m-%d %H:%M') if t.closed_at else '',
        ]
        ws.append(row)
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = border
            ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center")
        row_idx += 1

    _autofit(ws)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="Support_Report_{datetime.date.today()}.xlsx"'
    )
    return response


def generate_support_report_csv(queryset=None):
    """Generate a CSV IT support ticket report."""
    if queryset is None:
        queryset = SupportTicket.objects.select_related('employee', 'assigned_engineer').all()

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = (
        f'attachment; filename="Support_Report_{datetime.date.today()}.csv"'
    )
    writer = csv.writer(response)
    writer.writerow([
        "Ticket No.", "Employee", "Priority", "Category",
        "Status", "Subject", "Assigned Engineer", "Created At", "Closed At",
    ])
    for t in queryset:
        writer.writerow([
            t.ticket_number,
            f"{t.employee.first_name} {t.employee.last_name}",
            t.get_priority_display(),
            t.get_category_display(),
            t.get_status_display(),
            t.subject,
            f"{t.assigned_engineer.first_name} {t.assigned_engineer.last_name}" if t.assigned_engineer else '',
            t.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            t.closed_at.strftime('%Y-%m-%d %H:%M:%S') if t.closed_at else '',
        ])
    return response


# ─────────────────────────────────────────────
# Software License Report
# ─────────────────────────────────────────────

def generate_license_report_pdf(queryset=None):
    """Generate a PDF software license report."""
    if queryset is None:
        queryset = SoftwareLicense.objects.select_related('assigned_employee').all()

    buffer = io.BytesIO()
    doc, story = _pdf_doc(buffer, "Software License Report")

    headers = ["Software", "License Key", "Vendor", "Type",
               "Status", "Assigned To", "Purchase Date", "Expiry Date"]
    data = [headers]
    for lic in queryset:
        data.append([
            lic.software_name,
            lic.license_key[:20] + '...' if len(lic.license_key) > 20 else lic.license_key,
            lic.vendor or '-',
            lic.get_license_type_display(),
            lic.get_status_display(),
            (
                f"{lic.assigned_employee.first_name} {lic.assigned_employee.last_name}"
                if lic.assigned_employee else '-'
            ),
            str(lic.purchase_date) if lic.purchase_date else '-',
            str(lic.expiry_date) if lic.expiry_date else '-',
        ])

    tbl = Table(data, repeatRows=1)
    tbl.setStyle(_table_style())
    story.append(tbl)
    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="License_Report_{datetime.date.today()}.pdf"'
    )
    return response


def generate_license_report_excel(queryset=None):
    """Generate an Excel software license report."""
    if queryset is None:
        queryset = SoftwareLicense.objects.select_related('assigned_employee').all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Software Licenses"
    headers = [
        "Software", "License Key", "Vendor", "Type",
        "Status", "Assigned Employee", "Purchase Date", "Expiry Date",
    ]
    row_idx = _excel_header(ws, "Enterprise HRMS – Software License Report", headers)
    border = _thin_border()

    for lic in queryset:
        row = [
            lic.software_name,
            lic.license_key,
            lic.vendor or '',
            lic.get_license_type_display(),
            lic.get_status_display(),
            (
                f"{lic.assigned_employee.first_name} {lic.assigned_employee.last_name}"
                if lic.assigned_employee else ''
            ),
            str(lic.purchase_date) if lic.purchase_date else '',
            str(lic.expiry_date) if lic.expiry_date else '',
        ]
        ws.append(row)
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = border
            ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center")
        row_idx += 1

    _autofit(ws)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="License_Report_{datetime.date.today()}.xlsx"'
    )
    return response


def generate_license_report_csv(queryset=None):
    """Generate a CSV software license report."""
    if queryset is None:
        queryset = SoftwareLicense.objects.select_related('assigned_employee').all()

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = (
        f'attachment; filename="License_Report_{datetime.date.today()}.csv"'
    )
    writer = csv.writer(response)
    writer.writerow([
        "Software", "License Key", "Vendor", "Type",
        "Status", "Assigned Employee", "Purchase Date", "Expiry Date",
    ])
    for lic in queryset:
        writer.writerow([
            lic.software_name,
            lic.license_key,
            lic.vendor or '',
            lic.get_license_type_display(),
            lic.get_status_display(),
            (
                f"{lic.assigned_employee.first_name} {lic.assigned_employee.last_name}"
                if lic.assigned_employee else ''
            ),
            lic.purchase_date or '',
            lic.expiry_date or '',
        ])
    return response
