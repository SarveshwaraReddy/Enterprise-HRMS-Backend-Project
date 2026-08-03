import csv
import io
from decimal import Decimal
from django.db.models import Sum, Avg, Max, Min, Count
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from .models import PayrollRun, Payslip, SalaryStructure
from enterprise_hrms.employees.models import Employee
from enterprise_hrms.departments.models import Department


def get_payroll_summary_report(payroll_run_id=None, month=None, year=None):
    """
    Module 8: Payroll Summary Report
    Total Employees Paid, Gross Payroll, Total Deductions, Net Payroll
    """
    payslips = Payslip.objects.all()
    if payroll_run_id:
        payslips = payslips.filter(payroll_run_id=payroll_run_id)
    elif month and year:
        payslips = payslips.filter(payroll_run__payroll_month=month, payroll_run__payroll_year=year)

    aggregates = payslips.aggregate(
        total_employees=Count('employee', distinct=True),
        gross_payroll=Sum('gross_salary'),
        total_deductions=Sum('total_deductions'),
        net_payroll=Sum('net_salary')
    )

    return {
        'total_employees_paid': aggregates['total_employees'] or 0,
        'gross_payroll': aggregates['gross_payroll'] or Decimal('0.00'),
        'total_deductions': aggregates['total_deductions'] or Decimal('0.00'),
        'net_payroll': aggregates['net_payroll'] or Decimal('0.00'),
    }


def get_department_payroll_report(month=None, year=None):
    """
    Module 8: Department Payroll Report
    Payroll by Department, Average Salary, Highest Salary, Lowest Salary
    """
    payslips = Payslip.objects.all()
    if month and year:
        payslips = payslips.filter(payroll_run__payroll_month=month, payroll_run__payroll_year=year)

    departments = Department.objects.all()
    dept_reports = []

    for dept in departments:
        dept_payslips = payslips.filter(employee__department=dept)
        if dept_payslips.exists():
            stats = dept_payslips.aggregate(
                total_payroll=Sum('net_salary'),
                employee_count=Count('employee', distinct=True),
                avg_salary=Avg('net_salary'),
                max_salary=Max('net_salary'),
                min_salary=Min('net_salary')
            )
            dept_reports.append({
                'department_id': dept.id,
                'department_name': dept.name,
                'employee_count': stats['employee_count'] or 0,
                'total_payroll': stats['total_payroll'] or Decimal('0.00'),
                'avg_salary': round(stats['avg_salary'] or Decimal('0.00'), 2),
                'highest_salary': stats['max_salary'] or Decimal('0.00'),
                'lowest_salary': stats['min_salary'] or Decimal('0.00'),
            })
        else:
            dept_reports.append({
                'department_id': dept.id,
                'department_name': dept.name,
                'employee_count': 0,
                'total_payroll': Decimal('0.00'),
                'avg_salary': Decimal('0.00'),
                'highest_salary': Decimal('0.00'),
                'lowest_salary': Decimal('0.00'),
            })

    return dept_reports


def get_employee_salary_history(employee_id):
    """
    Module 8: Employee Salary History
    Display salary history by month for a given employee.
    """
    payslips = Payslip.objects.filter(employee_id=employee_id).order_by('-payroll_run__payroll_year', '-payroll_run__payroll_month')
    history = []
    for slip in payslips:
        history.append({
            'payslip_id': slip.id,
            'month': slip.payroll_run.payroll_month,
            'year': slip.payroll_run.payroll_year,
            'status': slip.payroll_run.status,
            'gross_salary': slip.gross_salary,
            'total_deductions': slip.total_deductions,
            'net_salary': slip.net_salary,
            'working_days': slip.working_days,
            'present_days': slip.present_days,
            'generated_at': slip.generated_at
        })
    return history


# Module 9: Export Reports (PDF, Excel, CSV)

def export_payroll_report_pdf(month, year):
    """
    Exports Monthly Payroll Report as PDF.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()

    story = []
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#1E3A8A'))
    story.append(Paragraph(f"Monthly Payroll Report - {month:02d}/{year}", title_style))
    story.append(Spacer(1, 10))

    summary = get_payroll_summary_report(month=month, year=year)
    story.append(Paragraph(f"<b>Total Employees Paid:</b> {summary['total_employees_paid']} | "
                           f"<b>Gross:</b> ${summary['gross_payroll']:,.2f} | "
                           f"<b>Deductions:</b> ${summary['total_deductions']:,.2f} | "
                           f"<b>Net Payroll:</b> ${summary['net_payroll']:,.2f}", styles['Normal']))
    story.append(Spacer(1, 15))

    payslips = Payslip.objects.filter(payroll_run__payroll_month=month, payroll_run__payroll_year=year)
    data = [["Employee ID", "Name", "Department", "Gross", "Deductions", "Net Salary"]]
    for p in payslips:
        emp = p.employee
        dept_name = emp.department.name if emp.department else "N/A"
        data.append([
            emp.employee_id,
            f"{emp.first_name} {emp.last_name}",
            dept_name,
            f"${p.gross_salary:,.2f}",
            f"${p.total_deductions:,.2f}",
            f"${p.net_salary:,.2f}"
        ])

    table = Table(data, colWidths=[80, 140, 110, 75, 75, 75])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
        ('PADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(table)
    doc.build(story)
    pdf_content = buffer.getvalue()
    buffer.close()
    return pdf_content


def export_payroll_register_excel(month=None, year=None):
    """
    Exports Payroll Register as Excel (.xlsx) workbook using openpyxl.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll Register"

    # Styling
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    headers = [
        "Employee ID", "Employee Name", "Department", "Designation",
        "Month/Year", "Working Days", "Present Days", "Gross Salary",
        "Total Deductions", "Net Salary", "Status"
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    payslips = Payslip.objects.all()
    if month and year:
        payslips = payslips.filter(payroll_run__payroll_month=month, payroll_run__payroll_year=year)

    for p in payslips:
        emp = p.employee
        dept_name = emp.department.name if emp.department else "N/A"
        row = [
            emp.employee_id,
            f"{emp.first_name} {emp.last_name}",
            dept_name,
            emp.designation,
            f"{p.payroll_run.payroll_month}/{p.payroll_run.payroll_year}",
            p.working_days,
            float(p.present_days),
            float(p.gross_salary),
            float(p.total_deductions),
            float(p.net_salary),
            p.payroll_run.status.capitalize()
        ]
        ws.append(row)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border

    buffer = io.BytesIO()
    wb.save(buffer)
    excel_content = buffer.getvalue()
    buffer.close()
    return excel_content


def export_payroll_transactions_csv(month=None, year=None):
    """
    Exports Payroll Transactions as CSV format.
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    headers = [
        "Payslip ID", "Employee ID", "Employee Name", "Email", "Department",
        "Month", "Year", "Gross Salary", "Total Deductions", "Net Salary",
        "Working Days", "Present Days", "Leave Days", "Generated At"
    ]
    writer.writerow(headers)

    payslips = Payslip.objects.all()
    if month and year:
        payslips = payslips.filter(payroll_run__payroll_month=month, payroll_run__payroll_year=year)

    for p in payslips:
        emp = p.employee
        dept_name = emp.department.name if emp.department else "N/A"
        writer.writerow([
            p.id,
            emp.employee_id,
            f"{emp.first_name} {emp.last_name}",
            emp.email,
            dept_name,
            p.payroll_run.payroll_month,
            p.payroll_run.payroll_year,
            f"{p.gross_salary:.2f}",
            f"{p.total_deductions:.2f}",
            f"{p.net_salary:.2f}",
            p.working_days,
            p.present_days,
            p.leave_days,
            p.generated_at.strftime('%Y-%m-%d %H:%M:%S')
        ])

    csv_content = buffer.getvalue()
    buffer.close()
    return csv_content
